import time
import tkinter as tk
from tkinter import ttk
from tkinter import Toplevel, Label, Entry, Button, messagebox
import tkinter.messagebox as mbox
from tkinter import filedialog
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException, TimeoutException
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import datetime
import openpyxl
import re
import sys
import os
import csv

# --- Biến driver toàn cục ---
browser = None
delete_buttons = []
dang_xoa_hs_trung = False
dong_test_hien_tai = None  # lưu dòng hiện tại khi test
dang_xoa_hs_7980 = False
dong_hien_tai_7980 = None  # Dòng hiện tại để duyệt danh sách 7980
ws_excel_7980 = None



# --- Hàm khởi động Chrome và điền thông tin tự động ---
def launchBrowser():
    chrome_options = Options()
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-background-timer-throttling")
    chrome_options.add_argument("--disable-backgrounding-occluded-windows")
    chrome_options.add_argument("--disable-renderer-backgrounding")

    browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    browser.get("https://gdbhyt.baohiemxahoi.gov.vn/")

    # ✅ Lấy thông tin từ file login.csv
    ma_cs, username, password = doc_thong_tin_dang_nhap()

    try:
        WebDriverWait(browser, 15).until(
            EC.presence_of_element_located((By.ID, "macskcb"))
        ).send_keys(ma_cs)

        WebDriverWait(browser, 10).until(
            EC.presence_of_element_located((By.ID, "username"))
        ).send_keys(username)

        WebDriverWait(browser, 10).until(
            EC.presence_of_element_located((By.ID, "password"))
        ).send_keys(password)

        WebDriverWait(browser, 5).until(
            EC.presence_of_element_located((By.ID, "Captcha_TB_I"))
        )
        browser.execute_script("document.getElementById('Captcha_TB_I').focus();")

    except TimeoutException:
        print("Không tìm thấy đủ các trường nhập sau 15 giây.")

    return browser


# Đọc thông tin từ file csv
def doc_thong_tin_dang_nhap():
    filepath = get_login_file_path()
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                return (
                    row.get("ma_co_so", ""),
                    row.get("ten_dang_nhap", ""),
                    row.get("mat_khau", "")
                )
    return "", "", ""


# Hàm để tìm đường dẫn file login.csv tránh trường hợp báo lỗi khi chạy file exe
def get_login_file_path():
    import sys
    if getattr(sys, 'frozen', False):
        return os.path.join(os.path.dirname(sys.executable), "login.csv")
    else:
        return os.path.join(os.path.abspath("."), "login.csv")




# --- Giao diện GUI ---
def resource_path(relative_path):
    """Lấy đường dẫn thực khi chạy file EXE hoặc khi chạy script trực tiếp"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.abspath(relative_path)

root = tk.Tk()
root.iconbitmap(resource_path("icon.ico"))

root.title("Tự động xóa cổng bảo hiểm")
root.geometry("450x700")





# Hàm ghi log
def ghi_log(message):
    text_box.config(state='normal')
    text_box.insert(tk.END, message + "\n")
    text_box.see(tk.END)  # Tự động cuộn xuống dòng cuối
    text_box.config(state='disabled')

# --- Hàm mở trình duyệt ---
def mo_chrome():
    global browser

    btn_login.config(state="disabled")  # ❌ Tạm vô hiệu hóa nút
    status_label.config(text="Đang khởi động trình duyệt...", fg="blue")
    root.update()  # ✅ Cập nhật giao diện ngay lập tức

    try:
        browser = launchBrowser()
        if browser:
            status_label.config(text="✅ Hãy nhập Captcha và bấm 'Đăng nhập' thủ công.", fg="green")
        else:
            status_label.config(text="❌ Không thể khởi động trình duyệt.", fg="red")
    except WebDriverException as e:
        status_label.config(text="❌ Lỗi khi mở Chrome", fg="red")
    
    btn_login.config(state="normal")  # ✅ Bật lại nút sau khi xong




# --- Hàm chung để chọn menu và lấy danh sách hồ sơ ---
def lay_danh_sach_ho_so(menu_ids, combobox_ids, output_var_name, ten_ho_so):
    global browser
    if browser is None:
        status_label.config(text="⚠️ Chưa đăng nhập cổng bảo hiểm", fg="orange")
        return
    
    # 🟦 Thông báo đang xử lý
    btn_load_hs_trung.config(state="disabled")  # ❌ Tạm vô hiệu hóa nút
    btn_load_hs_7980.config(state="disabled")  # ❌ Tạm vô hiệu hóa nút
    status_label.config(text=f"⏳ Đang load danh sách hồ sơ {ten_ho_so}...", fg="blue")
    root.update()  # ✅ Cập nhật giao diện ngay lập tức

    try:
        # 1. Đóng popup phiên bản nếu có
        try:
            popup = browser.find_element(By.ID, "popupInfoVersion_PW-1")
            if popup.is_displayed():
                browser.find_element(By.ID, "popupInfoVersion_HCB-1").click()
                time.sleep(0.5)
        except:
            pass  # Không có popup thì bỏ qua

        # 2. Click lần lượt các menu theo ID
        for menu_id in menu_ids:
            WebDriverWait(browser, 10).until(
                EC.element_to_be_clickable((By.ID, menu_id))
            ).click()

        # 3. Chọn các giá trị trong combobox lọc
        for combo_id, item_id in combobox_ids:
            WebDriverWait(browser, 10).until(
                EC.element_to_be_clickable((By.ID, combo_id))
            ).click()
            WebDriverWait(browser, 5).until(
                EC.element_to_be_clickable((By.ID, item_id))
            ).click()



        # # 4. Chọn tháng từ Combobox giao diện người dùng
        # thang_chon = combo_thang.get()                     # Ví dụ: "Tháng 7"
        # so_thang = int(thang_chon.split()[-1])             # Lấy số 7
        # index_thang = so_thang - 1                         # Index = 6

        # WebDriverWait(browser, 10).until(
        #     EC.element_to_be_clickable((By.ID, "cbx_thang_I"))
        # ).click()
        # WebDriverWait(browser, 5).until(
        #     EC.element_to_be_clickable((By.ID, f"cbx_thang_DDD_L_LBI{index_thang}T0"))
        # ).click()



        # 4. Chọn tháng từ Combobox giao diện người dùng
        thang_chon = combo_thang.get()                     # Ví dụ: "Tháng 7"
        so_thang = int(thang_chon.split()[-1])             # Lấy số 7
        index_thang = so_thang - 1                         # Index = 6

        # --- Kiểm tra nếu đã chọn đúng tháng ---
        try:
            selected_thang = browser.find_element(By.ID, "cbx_thang_I").get_attribute("value")
            if str(so_thang) in selected_thang:
                print(f"✅ Tháng {so_thang} đã được chọn sẵn, bỏ qua bước chọn.")
            else:
                # Nếu chưa chọn đúng thì mới thực hiện chọn
                WebDriverWait(browser, 10).until(
                    EC.element_to_be_clickable((By.ID, "cbx_thang_I"))
                ).click()

                # Đợi dropdown hiển thị ổn định
                WebDriverWait(browser, 5).until(
                    EC.visibility_of_element_located((By.ID, f"cbx_thang_DDD_L_LBI{index_thang}T0"))
                )

                # Thử click vài lần nếu lần đầu không hiệu lực
                for _ in range(3):
                    try:
                        browser.find_element(By.ID, f"cbx_thang_DDD_L_LBI{index_thang}T0").click()
                        break
                    except Exception:
                        time.sleep(0.3)

                # Xác nhận lại tháng sau khi chọn (không bắt buộc)
                selected_again = browser.find_element(By.ID, "cbx_thang_I").get_attribute("value")
                if str(so_thang) not in selected_again:
                    print("⚠️ Cảnh báo: Tháng chưa được chọn đúng.")
        except Exception as e:
            ghi_log(f"❌ Lỗi khi kiểm tra/chọn tháng: {e}")



        # 5. Bấm nút Tìm kiếm
        WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((By.ID, "bt_TimKiem"))
        ).click()

        # 6. Lấy kết quả tổng số hồ sơ
        try:
            summary_element = WebDriverWait(browser, 15).until(
                EC.visibility_of_element_located((By.CLASS_NAME, "dxp-summary"))
            )
            summary_text = summary_element.text  # Ví dụ: "Page 1 of 18 (348 items)"

            match = re.search(r"\((\d+)\s+items\)", summary_text)
            count = int(match.group(1)) if match else 0

            # 7. Gán biến toàn cục theo tên output_var_name
            globals()[output_var_name] = count

            # 8. Ghi log kết quả
            ghi_log(f"✅ Đã tìm thấy {count} hồ sơ {ten_ho_so}")
            status_label.config(text=f"✅ Đã tìm thấy {count} hồ sơ {ten_ho_so}", fg="green")

        except TimeoutException:
            ghi_log("❌ Lỗi tải trang, không tìm thấy kết quả sau 15s.")
            status_label.config(text="❌ Lỗi tải trang, không tìm thấy kết quả sau 15s", fg="red")
            btn_load_hs_trung.config(state="normal")
            btn_load_hs_7980.config(state="normal")
            return

    except Exception as e:
        status_label.config(text="❌ Lỗi tải trang", fg="red")
        ghi_log(f"❌ Lỗi tải trang: {e}")

    btn_load_hs_trung.config(state="normal")
    btn_load_hs_7980.config(state="normal")




# Load hồ sơ trùng
def load_ho_so_trung():
    menu_ids = [
        "HeaderMenu_DXI2_T",         # 🧭 Menu cấp 1: "Hồ sơ đề nghị thanh toán"
        "HeaderMenu_DXI2i0_T",       # 📄 Menu cấp 2: "Hồ sơ XML"
        "HeaderMenu_DXI2i0i2_T",     # 📋 Menu cấp 3: "Danh sách đề nghị thanh toán"
    ]
    combobox_ids = [
        ("cb_TrangThaiHS_I", "cb_TrangThaiHS_DDD_L_LBI2T0"),  # ✅ Trạng thái hồ sơ: "Hồ sơ trùng"
        ("cb_TrangThaiTT_I", "cb_TrangThaiTT_DDD_L_LBI0T0"),  # 💰 Trạng thái thanh toán: "Tất cả"
    ]
    lay_danh_sach_ho_so(menu_ids, combobox_ids, "so_ho_so_trung", "trùng")



# Load hồ sơ 7980
def load_ho_so_7980():
    menu_ids = [
        "HeaderMenu_DXI2_T",         # 🧭 Menu cấp 1: "Hồ sơ đề nghị thanh toán"
        "HeaderMenu_DXI2i1_T",       # 📁 Menu cấp 2: "Hồ sơ 7980"
        "HeaderMenu_DXI2i1i2_",     # 📋 Menu cấp 3: "Danh sách đề nghị thanh toán 7980a"
    ]
    combobox_ids = [
        ("cb_TrangThaiHS_I", "cb_TrangThaiHS_DDD_L_LBI0T0"),  # ✅ Trạng thái hồ sơ: "Tất cả"
        ("cb_TrangThaiTT_I", "cb_TrangThaiTT_DDD_L_LBI0T0"),  # 💰 Trạng thái thanh toán: "Tất cả"
    ]
    lay_danh_sach_ho_so(menu_ids, combobox_ids, "so_ho_so_7980", "79/80")








# Xóa hồ sơ trùng
def xoa_danh_sach_ho_so_trung():
    global dang_xoa_hs_trung

    def dem_so_ho_so_tren_trang():
        try:
            summary_element = WebDriverWait(browser, 5).until(
                EC.visibility_of_element_located((By.CLASS_NAME, "dxp-summary"))
            )
            summary_text = summary_element.text  # Ví dụ: "Page 1 of 18 (348 items)"
            match = re.search(r"\((\d+)\s+items\)", summary_text)
            return int(match.group(1)) if match else 0
        except:
            return -1  # Không còn thấy phần tử nữa

    if not dang_xoa_hs_trung:
        return

    def xoa_tiep():
        global dang_xoa_hs_trung

        if not dang_xoa_hs_trung:
            ghi_log("⏹️ Tạm dừng")
            status_label.config(text="⏹️ Tạm dừng", fg="red")
            btn_delete_hs_trung.config(text="Xóa HS Trùng")
            return

        current_count = dem_so_ho_so_tren_trang()

        if current_count == -1:
            # Cho phép thử tiếp nếu chỉ mất tạm thời
            current_count = 0

        if current_count == 0:
            ghi_log("✅ Không tìm thấy hồ sơ trùng.")
            status_label.config(text="✅ Không tìm thấy hồ sơ trùng.", fg="green")
            btn_delete_hs_trung.config(text="Xóa HS Trùng")
            dang_xoa_hs_trung = False
            return

        try:
            row = WebDriverWait(browser, 5).until(
                EC.presence_of_element_located((By.XPATH, "//tr[contains(@id, 'gvDanhSachHoSo_DXDataRow')]"))
            )
            cols = row.find_elements(By.TAG_NAME, "td")
            ho_ten = cols[8].text.strip() if len(cols) > 8 else "Không rõ tên"
            icon_xoa = cols[26].find_element(By.TAG_NAME, "img")

            browser.execute_script("arguments[0].click();", icon_xoa)

            WebDriverWait(browser, 10).until(
                EC.visibility_of_element_located((By.ID, "PopupThongBaoXoa_PWH-1"))
            )
            btn_co = WebDriverWait(browser, 3).until(
                EC.element_to_be_clickable((By.ID, "btnCo"))
            )
            btn_co.click()

            # Theo dõi số lượng thay đổi
            fail_count = 0
            for _ in range(20):  # tối đa 10s
                time.sleep(0.5)
                new_count = dem_so_ho_so_tren_trang()

                print(f"[DEBUG] current_count = {current_count}, new_count = {new_count}")

                if new_count == -1:
                    fail_count += 1
                    if fail_count >= 3:
                        ghi_log("⚠️ Không tìm thấy bảng kết quả sau nhiều lần. Dừng lại.")
                        status_label.config(text="⚠️ Không tìm thấy bảng kết quả sau nhiều lần.", fg="red")
                        btn_delete_hs_trung.config(text="Xóa HS Trùng")
                        dang_xoa_hs_trung = False
                        return
                    continue  # thử lại

                fail_count = 0  # reset nếu bình thường

                if new_count < current_count:
                    print(f"[DEBUG] Hồ sơ giảm từ {current_count} → {new_count}")
                    # Đóng popup nếu có
                    try:
                        WebDriverWait(browser, 5).until(
                            EC.visibility_of_element_located((By.ID, "popup_message_PW-1"))
                        )
                        btn_close = WebDriverWait(browser, 3).until(
                            EC.element_to_be_clickable((By.ID, "popup_message_HCB-1"))
                        )
                        btn_close.click()
                    except:
                        pass

                    ghi_log(f"🗑️ Đã xóa hồ sơ: {ho_ten}")
                    break
            else:
                ghi_log(f"❌ Xóa hồ sơ {ho_ten} thất bại (số lượng không đổi)")
                btn_delete_hs_trung.config(text="Xóa HS Trùng")
                dang_xoa_hs_trung = False
                return

        except Exception as e:
            ghi_log(f"❌ Lỗi khi xóa hồ sơ: {e}")

        root.after(500, xoa_tiep)

    xoa_tiep()








def toggle_xoa_ho_so_trung():
    global dang_xoa_hs_trung

    if not dang_xoa_hs_trung:
        # Bắt đầu xóa
        dang_xoa_hs_trung = True
        btn_delete_hs_trung.config(text="⏹️ Dừng")
        xoa_danh_sach_ho_so_trung()
    else:
        # Dừng thao tác xóa
        dang_xoa_hs_trung = False
        btn_delete_hs_trung.config(text="Xóa HS Trùng")








# Hàm để chọn file Excel
def chon_file_excel():
    global dong_test_hien_tai
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx *.xls")],
        title="Chọn file Excel"
    )
    if file_path:
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, file_path)
        dong_test_hien_tai = None  # Reset dòng test
        ghi_log(f"📂 Đã chọn file: {file_path}")

# Test file excel
def test_in_thong_tin_excel():
    global dong_test_hien_tai

    ma_the_col = combo_mt.get().strip().upper()
    ho_ten_col = combo_ht.get().strip().upper()
    ngay_vao_col = combo_nv.get().strip().upper()
    ngay_ra_col = combo_nr.get().strip().upper()

    file_path = entry_file_path.get().strip()
    if not file_path:
        ghi_log("❌ Chưa chọn file Excel.")
        return

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        ghi_log(f"❌ Lỗi khi mở file Excel: {e}")
        return

    try:
        start = int(entry_start.get().strip())
        end = int(entry_end.get().strip())
    except ValueError:
        ghi_log("⚠️ Vui lòng nhập số nguyên cho dòng bắt đầu và kết thúc.")
        return

    if dong_test_hien_tai is None:
        dong_test_hien_tai = start

    if dong_test_hien_tai > end:
        ghi_log("✅ Đã test đến dòng cuối cùng.")
        return

    try:
        ma_the_val = ws[f"{ma_the_col}{dong_test_hien_tai}"].value
        ho_ten_val = ws[f"{ho_ten_col}{dong_test_hien_tai}"].value
        ngay_vao_val = ws[f"{ngay_vao_col}{dong_test_hien_tai}"].value
        ngay_ra_val = ws[f"{ngay_ra_col}{dong_test_hien_tai}"].value

        # Gộp thông tin thành 1 dòng log
        log_line = f"{dong_test_hien_tai}: {ma_the_val} | {ho_ten_val} | {ngay_vao_val} | {ngay_ra_val}"
        ghi_log(log_line)

    except Exception as e:
        ghi_log(f"❌ Lỗi khi đọc dòng {dong_test_hien_tai}: {e}")

    dong_test_hien_tai += 1


def mo_file_excel_7980():
    global ws_excel_7980

    file_path = entry_file_path.get().strip()
    if not file_path:
        ghi_log("❌ Chưa chọn file Excel.")
        return False

    try:
        wb = openpyxl.load_workbook(file_path)
        ws_excel_7980 = wb.active
        return True
    except Exception as e:
        ghi_log(f"❌ Lỗi khi mở file Excel: {e}")
        status_label.config(text="❌ Lỗi khi mở file Excel", fg="red")
        return False

# Duyệt qua danh sách excel, tìm kiếm, so sánh và xóa hồ sơ 7980
def xoa_ho_so_7980():
    global dong_hien_tai, dang_xoa_hs_7980

    if not dang_xoa_hs_7980:
        btn_delete_hs_7980.config(text="⏹️ Dừng xóa")
        ghi_log("🚀 Bắt đầu xóa hồ sơ 79/80...")
        status_label.config(text="🚀 Đang xóa hồ sơ 79/80...", fg="blue")
        dong_hien_tai = None
        dang_xoa_hs_7980 = True

        if mo_file_excel_7980():
            xoa_tiep_dong_7980()
        else:
            dang_xoa_hs_7980 = False
            btn_delete_hs_7980.config(text="Xóa HS 79/80")
    else:
        dang_xoa_hs_7980 = False
        btn_delete_hs_7980.config(text="Xóa HS 79/80")
        ghi_log("⏹️ Đã dừng xóa.")
        status_label.config(text="⏹️ Đã dừng xóa.", fg="red")



def xoa_tiep_dong_7980():
    import time
    global dong_hien_tai, dang_xoa_hs_7980

    if not dang_xoa_hs_7980:
        return

    # --- Đọc dữ liệu từ file Excel ---
    file_path = entry_file_path.get().strip()
    if not file_path:
        ghi_log("❌ Chưa chọn file Excel.")
        status_label.config(text="❌ Chưa chọn file Excel.", fg="red")
        return

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        ghi_log(f"❌ Lỗi khi mở file Excel: {e}")
        status_label.config(text="❌ Lỗi khi mở file Excel", fg="red")
        return

    try:
        start = int(entry_start.get().strip())
        end = int(entry_end.get().strip())
    except ValueError:
        ghi_log("⚠️ Vui lòng nhập số nguyên cho dòng bắt đầu và kết thúc.")
        status_label.config(text="⚠️ Vui lòng nhập số nguyên cho dòng bắt đầu và kết thúc.", fg="red")
        return

    if dong_hien_tai is None:
        dong_hien_tai = start

    if dong_hien_tai > end:
        ghi_log("✅ Đã duyệt hết tất cả các dòng.")
        status_label.config(text="✅ Đã duyệt hết tất cả các dòng.", fg="blue")
        btn_delete_hs_7980.config(text="Xóa HS 79/80")
        dang_xoa_hs_7980 = False
        return

    try:
        # --- Hàm phụ đếm số dòng kết quả ---
        def dem_so_dong_ket_qua():
            return len(browser.find_elements(By.XPATH, "//tr[starts-with(@id, 'gvDanhSachBHYT7980_DXDataRow')]"))

        # --- Lấy dữ liệu từ dòng hiện tại ---
        ma_the_col = combo_mt.get().strip().upper()
        ho_ten_col = combo_ht.get().strip().upper()
        ngay_vao_col = combo_nv.get().strip().upper()
        ngay_ra_col = combo_nr.get().strip().upper()

        ma_the_val = ws[f"{ma_the_col}{dong_hien_tai}"].value
        ho_ten_val = str(ws[f"{ho_ten_col}{dong_hien_tai}"].value).strip()
        ngay_vao_val = str(ws[f"{ngay_vao_col}{dong_hien_tai}"].value).strip()
        ngay_ra_val = str(ws[f"{ngay_ra_col}{dong_hien_tai}"].value).strip()

        if not ma_the_val:
            ghi_log(f"{dong_hien_tai}: ⚠️ Không có mã thẻ. Bỏ qua.")
            dong_hien_tai += 1
            root.after(500, xoa_tiep_dong_7980)
            return

        # --- Tìm lại input mỗi lần để tránh stale element ---
        input_box = WebDriverWait(browser, 5).until(
            EC.presence_of_element_located((By.ID, "gvDanhSachBHYT7980_DXFREditorcol3_I"))
        )

        input_val = input_box.get_attribute("value").strip()

        if input_val:
            input_box.clear()

            # --- Chờ số dòng thay đổi sau khi clear ---
            old_count = dem_so_dong_ket_qua()
            start_time = time.time()
            while time.time() - start_time < 30:
                new_count = dem_so_dong_ket_qua()
                if new_count != old_count:
                    print(f"Đã lọc thành công {new_count} kết quả")
                    break
                time.sleep(0.5)
            else:
                ghi_log("❌ Lỗi tải trang, không lọc được kết quả")
                btn_delete_hs_7980.config(text="Xóa HS 79/80")
                dang_xoa_hs_7980 = False
                return

            # --- Tìm lại input box sau khi reload ---
            input_box = WebDriverWait(browser, 5).until(
                EC.presence_of_element_located((By.ID, "gvDanhSachBHYT7980_DXFREditorcol3_I"))
            )

        # --- Nhập mã thẻ ---
        input_box.send_keys(str(ma_the_val))

        # --- Chờ số dòng thay đổi sau khi nhập mã thẻ ---
        old_count = dem_so_dong_ket_qua()
        start_time = time.time()
        while time.time() - start_time < 10:
            new_count = dem_so_dong_ket_qua()
            if new_count != old_count:
                print(f"Đã lọc thành công {new_count} kết quả")
                break
            time.sleep(0.5)
        else:
            ghi_log("❌ Lỗi tải trang, không lọc được kết quả")
            btn_delete_hs_7980.config(text="Xóa HS 79/80")
            dang_xoa_hs_7980 = False
            return

        # --- Tìm danh sách kết quả ---
        rows = browser.find_elements(By.XPATH, "//tr[starts-with(@id, 'gvDanhSachBHYT7980_DXDataRow')]")
        found = False

        for row in rows:
            cols = row.find_elements(By.TAG_NAME, "td")
            if len(cols) >= 19:
                ho_ten = cols[5].text.strip()
                ngay_vao = cols[8].text.strip().replace(" SA", "").replace(" CH", "")
                ngay_ra = cols[9].text.strip().replace(" SA", "").replace(" CH", "")

                if ho_ten == ho_ten_val and ngay_vao == ngay_vao_val and ngay_ra == ngay_ra_val:
                    try:
                        cols = row.find_elements(By.TAG_NAME, "td")
                        delete_btn = cols[18].find_element(By.TAG_NAME, "input")

                        browser.execute_script("arguments[0].click();", delete_btn)

                        # ⏳ Chờ popup xác nhận xóa hiện ra
                        WebDriverWait(browser, 5).until(
                            EC.visibility_of_element_located((By.ID, "PopupThongBaoXoa_PWH-1"))
                        )

                        # ✅ Bấm nút "Có"
                        btn_co = WebDriverWait(browser, 3).until(
                            EC.element_to_be_clickable((By.ID, "btnCo_CD"))
                        )
                        btn_co.click()

                        # ⏳ Chờ popup thông báo kết quả xóa (có thể xuất hiện)
                        try:
                            WebDriverWait(browser, 5).until(
                                EC.visibility_of_element_located((By.ID, "popup_message_PWH-1"))
                            )
                            # ✅ Bấm nút Close để đóng popup
                            btn_close = WebDriverWait(browser, 3).until(
                                EC.element_to_be_clickable((By.ID, "popup_message_HCB-1"))
                            )
                            btn_close.click()
                        except:
                            ghi_log(f"{dong_hien_tai}: ❌ xóa {ho_ten} thất bại (không thấy popup xác nhận)")
                            dang_xoa_hs_7980 = False
                            btn_delete_hs_7980.config(text="Xóa HS 79/80")
                            return

                        ghi_log(f"{dong_hien_tai}: 🗑️ Đã xóa: {ho_ten}")
                        found = True
                        break

                    except Exception as e:
                        ghi_log(f"{dong_hien_tai}: ❌ Không thể xóa: {e}")


        if not found:
            ghi_log(f"{dong_hien_tai}: ❌ Không tìm thấy hồ sơ của {ho_ten}")

    except Exception as e:
        ghi_log(f"{dong_hien_tai}: ❌ Lỗi: {e}")

    dong_hien_tai += 1
    root.after(500, xoa_tiep_dong_7980)






# Cửa sổ Cài đặt: Đã giải quyết vấn đề hiển thị chính giữa cửa sổ cha và không bị nháy 2 lần
def mo_cai_dat():
    # --- Tính toán vị trí trước ---
    w, h = 300, 250
    root.update_idletasks()
    root_x = root.winfo_rootx()
    root_y = root.winfo_rooty()
    root_w = root.winfo_width()
    root_h = root.winfo_height()
    x = root_x + (root_w // 2) - (w // 2)
    y = root_y + (root_h // 2) - (h // 2)

    # --- Dùng after để tránh nháy ---
    def tao_cua_so():
        win = tk.Toplevel()
        win.withdraw()  # ✅ Ẩn ngay sau khi tạo để tránh nháy
        win.title("Cài đặt đăng nhập")
        win.iconbitmap(resource_path("icon.ico"))  # ✅ Icon cửa sổ
        win.geometry(f"{w}x{h}+{x}+{y}")
        win.resizable(False, False)
        win.transient(root)
        win.grab_set()

        # --- Giao diện ---
        tk.Label(win, text="Mã cơ sở KCB:", font=("Arial", 10)).pack(pady=(10, 0))
        entry_ma_cs = tk.Entry(win, font=("Arial", 10), justify="center")
        entry_ma_cs.pack()

        tk.Label(win, text="Tên đăng nhập:", font=("Arial", 10)).pack(pady=(10, 0))
        entry_username = tk.Entry(win, font=("Arial", 10), justify="center")
        entry_username.pack()

        tk.Label(win, text="Mật khẩu:", font=("Arial", 10)).pack(pady=(10, 0))
        entry_password = tk.Entry(win, font=("Arial", 10), justify="center")
        entry_password.pack()

        # --- Đọc dữ liệu từ CSV ---
        filepath = get_login_file_path()
        if os.path.exists(filepath):
            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    entry_ma_cs.insert(0, row.get("ma_co_so", ""))
                    entry_username.insert(0, row.get("ten_dang_nhap", ""))
                    entry_password.insert(0, row.get("mat_khau", ""))
                    break

        # --- Lưu ---
        def luu():
            ma_cs = entry_ma_cs.get().strip()
            username = entry_username.get().strip()
            password = entry_password.get().strip()

            if not ma_cs or not username or not password:
                messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập đầy đủ.")
                return

            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=["ma_co_so", "ten_dang_nhap", "mat_khau"])
                writer.writeheader()
                writer.writerow({
                    "ma_co_so": ma_cs,
                    "ten_dang_nhap": username,
                    "mat_khau": password
                })

            messagebox.showinfo("Thành công", "Đã lưu.")
            win.destroy()

        tk.Button(win, text="Lưu", font=("Arial", 10), width=10, command=luu).pack(pady=20)
        
        # --- Footer ---
        tk.Label(win, text="07/2025 - buitiencong@gmail.com", font=("Arial", 8), fg="gray").pack(side="bottom", pady=(5, 5))

        win.deiconify()  # ✅ Hiển thị lại sau khi setup xong

    root.after(1, tao_cua_so)






























# === Đăng nhập + Combobox tháng ===
# --- Dòng chứa nút Đăng nhập (giữa) và Cài đặt (phải) ---
top_button_frame = tk.Frame(root)
top_button_frame.pack(fill="x", pady=10, padx=10)

# Dùng lưới 3 cột: spacer trái - nút đăng nhập - spacer phải
top_button_frame.columnconfigure(0, weight=1)
top_button_frame.columnconfigure(1, weight=0)
top_button_frame.columnconfigure(2, weight=1)

btn_login = tk.Button(top_button_frame, text="Đăng nhập cổng bảo hiểm", font=("Arial", 12), command=mo_chrome)
btn_login.grid(row=0, column=1)

# Nút "Cài đặt" căn sát phải bằng place
btn_caidat = tk.Button(top_button_frame, text="⚙️", font=("Arial", 12), command=mo_cai_dat)
btn_caidat.place(relx=1.0, x=0, y=0, anchor="ne")



thang_hien_tai = datetime.datetime.now().month
index_mac_dinh = thang_hien_tai - 1

selected_thang = tk.StringVar()
combo_thang = ttk.Combobox(root, textvariable=selected_thang, font=("Arial", 11), width=10, state="readonly")
combo_thang['values'] = [f"Tháng {i}" for i in range(1, 13)]
combo_thang.current(index_mac_dinh)
combo_thang.pack(pady=5)

# === Notebook chứa 2 tab ===
style = ttk.Style()
style.configure("TNotebook.Tab", font=("Arial", 11), padding=[10, 5])  # Font mặc định
style.map("TNotebook.Tab",
    font=[("selected", ("Arial", 12, "bold"))],         # In đậm khi được chọn
    foreground=[("selected", "red")]                    # Màu chữ đỏ khi được chọn
)

# 👉 Bọc Notebook bằng Frame có padding trái/phải
notebook_frame = tk.Frame(root)
notebook_frame.pack(padx=10, pady=10, fill="x")  # bỏ expand để không bị kéo dãn theo Text box

# 👉 Notebook có chiều cao thay đổi được
notebook = ttk.Notebook(notebook_frame, height=300)  # height mặc định ban đầu
notebook.pack(fill="x")

# === Tab 1: Hồ sơ trùng ===
tab_hs_trung = tk.Frame(notebook)
notebook.add(tab_hs_trung, text="Hồ sơ trùng")

frame_buttons = tk.Frame(tab_hs_trung)  # 👉 Frame chứa các button, căn giữa theo cả 2 chiều
frame_buttons.pack(expand=True)

btn_load_hs_trung = tk.Button(frame_buttons, text="Load hồ sơ trùng", font=("Arial", 12), command=load_ho_so_trung)
btn_load_hs_trung.pack(pady=10)

btn_delete_hs_trung = tk.Button(frame_buttons, text="Xóa hồ sơ trùng", font=("Arial", 12), command=toggle_xoa_ho_so_trung)
btn_delete_hs_trung.pack(pady=10)

# === Tab 2: Hồ sơ 79/80 ===
tab_hs_7980 = tk.Frame(notebook)
notebook.add(tab_hs_7980, text="Hồ sơ 79/80")

btn_load_hs_7980 = tk.Button(tab_hs_7980, text="Load hồ sơ 79/80", font=("Arial", 12), command=load_ho_so_7980)
btn_load_hs_7980.pack(pady=10)

# === Sự kiện thay đổi chiều cao theo tab ===
def on_tab_change(event):
    status_label.config(text="", fg="black")

    selected_tab = event.widget.select()
    tab_text = notebook.tab(selected_tab, "text")

    if tab_text == "Hồ sơ trùng":
        notebook.configure(height=120)
    elif tab_text == "Hồ sơ 79/80":
        notebook.configure(height=200)

notebook.bind("<<NotebookTabChanged>>", on_tab_change)


# --- Dòng chọn file Excel ---
def set_placeholder(entry, text):
    entry.insert(0, text)
    entry.config(fg='gray')

    def on_focus_in(event):
        if entry.get() == text:
            entry.delete(0, tk.END)
            entry.config(fg='black')

    def on_focus_out(event):
        if not entry.get():
            entry.insert(0, text)
            entry.config(fg='gray')

    entry.bind("<FocusIn>", on_focus_in)
    entry.bind("<FocusOut>", on_focus_out)

def chon_file_excel():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        if entry_file_path.get() == "Chọn file excel đối chiếu":
            entry_file_path.delete(0, tk.END)
            entry_file_path.config(fg='black')
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, file_path)

file_select_frame = tk.Frame(tab_hs_7980)
file_select_frame.pack(pady=5)

entry_file_path = tk.Entry(file_select_frame, width=30, font=("Arial", 10))
entry_file_path.pack(side="left", padx=5)
set_placeholder(entry_file_path, "Chọn file excel đối chiếu")

btn_browse_file = tk.Button(file_select_frame, text="Chọn file", command=chon_file_excel)
btn_browse_file.pack(side="left")

# --- Dòng chọn cột ---
chu_cai_list = [chr(i) for i in range(65, 91)]

column_select_frame = tk.Frame(tab_hs_7980)
column_select_frame.pack(pady=5)

label_mt = tk.Label(column_select_frame, text="Mã thẻ")
label_mt.pack(side="left", padx=2)
combo_mt = ttk.Combobox(column_select_frame, values=chu_cai_list, width=3)
combo_mt.pack(side="left", padx=2)

label_ht = tk.Label(column_select_frame, text="Họ tên")
label_ht.pack(side="left", padx=2)
combo_ht = ttk.Combobox(column_select_frame, values=chu_cai_list, width=3)
combo_ht.pack(side="left", padx=2)

label_nv = tk.Label(column_select_frame, text="Ngày vào")
label_nv.pack(side="left", padx=2)
combo_nv = ttk.Combobox(column_select_frame, values=chu_cai_list, width=3)
combo_nv.pack(side="left", padx=2)

label_nr = tk.Label(column_select_frame, text="Ngày ra")
label_nr.pack(side="left", padx=2)
combo_nr = ttk.Combobox(column_select_frame, values=chu_cai_list, width=3)
combo_nr.pack(side="left", padx=2)

# --- Dòng nhập dòng bắt đầu/kết thúc ---
row_range_frame = tk.Frame(tab_hs_7980)
row_range_frame.pack(pady=5)

label_start = tk.Label(row_range_frame, text="Dòng bắt đầu")
label_start.pack(side="left", padx=2)
entry_start = tk.Entry(row_range_frame, width=6)
entry_start.pack(side="left", padx=2)

label_end = tk.Label(row_range_frame, text="Dòng kết thúc")
label_end.pack(side="left", padx=2)
entry_end = tk.Entry(row_range_frame, width=6)
entry_end.pack(side="left", padx=2)

btn_test_range = tk.Button(row_range_frame, text="Test", command=test_in_thong_tin_excel)
btn_test_range.pack(side="left", padx=5)

btn_delete_hs_7980 = tk.Button(tab_hs_7980, text="Xóa hồ sơ 79/80", font=("Arial", 12), command=xoa_ho_so_7980)
btn_delete_hs_7980.pack(pady=5)

# === Status và TextBox ===
status_label = tk.Label(root, text="", font=("Arial", 10))
status_label.pack()

text_box = tk.Text(root, height=25, font=("Arial", 10))
text_box.pack(padx=10, pady=10, fill="both")
text_box.config(state='disabled')

root.mainloop()
